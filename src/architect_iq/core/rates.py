"""Rate cards: loadable roles-and-rates files for leverage modeling (spec §2.6).

A rate card is a set of (discipline, tier, location) -> day_rate rows. Loading a
different card lets you model a different leverage model (seniority/location mix)
and reprice an estimate without re-running the effort math. Files load as CSV,
XLSX, or YAML (the pricing schema in data/SCHEMA.md).
"""

from __future__ import annotations

import io

from ..models.enums import Location
from ..models.results import MonteCarloResult, Percentiles
from ..models.solution_graph import SolutionGraph
from ..models.team import RateRow

_RATE_HEADERS = {"day_rate", "dayrate", "rate", "day rate"}


class RateCard:
    """Resolves a day rate by (discipline, tier, location), with a US fallback."""

    def __init__(self, rows: list[RateRow]):
        self.rows = rows
        self._index = {(r.discipline, r.tier, r.location): r.day_rate for r in rows}

    def rate_for(self, discipline: str, tier: str, location: Location) -> float:
        return (
            self._index.get((discipline, tier, location))
            or self._index.get((discipline, tier, Location.US))
            or 0.0
        )

    def blended_rate(self, discipline: str, tier: str, mix: dict[str, float]) -> float:
        """Weighted day rate across a location mix, e.g. {'US': 0.6, 'NS': 0.4}."""
        if not mix:
            return self.rate_for(discipline, tier, Location.US)
        total = 0.0
        weight = 0.0
        for loc, w in mix.items():
            try:
                location = Location(str(loc).upper())
            except ValueError:
                continue
            rate = self._index.get((discipline, tier, location)) or self._index.get((discipline, tier, Location.US)) or 0.0
            total += rate * w
            weight += w
        return total / weight if weight else 0.0

    def summary(self) -> dict:
        return {
            "rows": len(self.rows),
            "disciplines": sorted({r.discipline for r in self.rows}),
            "tiers": sorted({r.tier for r in self.rows}),
            "locations": sorted({r.location.value for r in self.rows}),
        }


def _row_from_mapping(m: dict) -> RateRow | None:
    """Build a RateRow from a case-insensitive header mapping."""
    norm = {str(k).strip().lower(): v for k, v in m.items() if k is not None}
    discipline = norm.get("discipline")
    tier = norm.get("tier")
    location = norm.get("location")
    rate = next((norm[h] for h in _RATE_HEADERS if h in norm), None)
    if not discipline or not tier or location is None or rate in (None, ""):
        return None
    return RateRow(
        discipline=str(discipline).strip(),
        tier=str(tier).strip(),
        location=Location(str(location).strip().upper()),
        day_rate=float(rate),
    )


def parse_rate_file(raw: bytes, filename: str) -> list[RateRow]:
    """Parse a rates file (.csv/.xlsx/.yaml). Raises ValueError with a clear
    message on bad input or missing columns."""
    lower = filename.lower()
    if lower.endswith(".csv"):
        rows = _parse_csv(raw)
    elif lower.endswith(".xlsx"):
        rows = _parse_xlsx(raw)
    elif lower.endswith((".yaml", ".yml")):
        rows = _parse_yaml(raw)
    else:
        raise ValueError("rates file must be .csv, .xlsx, or .yaml")
    if not rows:
        raise ValueError(
            "no rate rows parsed; expected columns: discipline, tier, location, day_rate"
        )
    return rows


def _parse_csv(raw: bytes) -> list[RateRow]:
    import csv

    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    return [r for r in (_row_from_mapping(d) for d in reader) if r]


def _parse_xlsx(raw: bytes) -> list[RateRow]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c).strip().lower() if c is not None else "" for c in next(rows_iter, [])]
    out: list[RateRow] = []
    for row in rows_iter:
        mapping = dict(zip(header, row))
        parsed = _row_from_mapping(mapping)
        if parsed:
            out.append(parsed)
    wb.close()
    return out


def _parse_yaml(raw: bytes) -> list[RateRow]:
    import yaml

    data = yaml.safe_load(raw.decode("utf-8", errors="replace")) or {}
    return [r for r in (_row_from_mapping(d) for d in data.get("rates", [])) if r]


def recost(graph: SolutionGraph, card: RateCard) -> SolutionGraph:
    """Reprice an estimate under a new rate card.

    Effort and duration are held fixed (rates don't change the work); only role
    day rates, monthly/total cost, and the cost distribution move. Team size and
    velocity are unchanged, so duration is unchanged.
    """
    variables = graph.variables
    updated = graph.model_copy(deep=True)

    for role in updated.team_plan.roles:
        role.day_rate = card.rate_for(role.discipline, role.tier, role.location)

    monthly = sum(r.day_rate * r.allocated * variables.working_month_days for r in updated.team_plan.roles)

    bottom_up = graph.deterministic.total_points if graph.deterministic else 0.0
    velocity = 0.0
    if graph.deterministic and graph.deterministic.per_phase_velocity:
        velocity = next(iter(graph.deterministic.per_phase_velocity.values()), 0.0)
    duration_sprints = bottom_up / velocity if velocity else 0.0
    duration_months = duration_sprints * variables.weeks_in_sprint / 4.345
    total_cost = round(monthly * duration_months, 2)

    updated.team_plan.monthly_cost = round(monthly, 2)
    updated.team_plan.total_cost = total_cost
    if updated.deterministic:
        updated.deterministic.total_cost = total_cost

    # Cost scales linearly with effort points; reprice the cost percentiles from
    # the (unchanged) effort percentiles.
    if updated.monte_carlo and bottom_up:
        cost_per_point = total_cost / bottom_up
        eff = updated.monte_carlo.effort_points
        updated.monte_carlo = MonteCarloResult(
            iterations=updated.monte_carlo.iterations,
            effort_points=eff,
            duration_sprints=updated.monte_carlo.duration_sprints,
            cost=Percentiles(
                p10=round(eff.p10 * cost_per_point, 2),
                p50=round(eff.p50 * cost_per_point, 2),
                p80=round(eff.p80 * cost_per_point, 2),
                p90=round(eff.p90 * cost_per_point, 2),
            ),
        )

    updated.assumptions = list(updated.assumptions) + [
        f"Re-costed with a custom rate card ({card.summary()['rows']} rows)."
    ]
    return updated
